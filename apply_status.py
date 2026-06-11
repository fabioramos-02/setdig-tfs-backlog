"""
Aplica status_updates.json (exportado pelo backlog_viewer.html) no backlog.xlsx
e regenera JSON + HTML.

Uso: python apply_status.py [caminho\\status_updates.json]
     (default: %USERPROFILE%\\Downloads\\status_updates.json)
"""
import json
import os
import subprocess
import sys

import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(BASE, "backlog.xlsx")
STATUSES = ("A fazer", "Em progresso", "Finalizado")
DEFAULT_JSON = os.path.join(os.path.expanduser("~"), "Downloads", "status_updates.json")


def main():
    json_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_JSON
    if not os.path.exists(json_path):
        print(f"[ERRO] Arquivo não encontrado: {json_path}")
        sys.exit(1)

    with open(json_path, encoding="utf-8") as f:
        updates = json.load(f)
    if not updates:
        print("[OK] Nenhum override no JSON — nada a aplicar.")
        return

    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Backlog"]
    headers = [c.value for c in ws[1]]
    if "status" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="status")
        headers.append("status")
    id_col = headers.index("csv_id") + 1
    st_col = headers.index("status") + 1

    applied, skipped = [], []
    rows_by_id = {}
    for row in ws.iter_rows(min_row=2):
        v = row[id_col - 1].value
        if v is not None:
            rows_by_id[str(int(v))] = row[0].row

    for cid, st in updates.items():
        if st not in STATUSES:
            skipped.append(f"{cid} (status inválido: {st!r})")
            continue
        r = rows_by_id.get(str(cid))
        if r is None:
            skipped.append(f"{cid} (csv_id não encontrado)")
            continue
        ws.cell(row=r, column=st_col, value=st)
        applied.append(f"{cid} -> {st}")

    wb.save(XLSX)
    for a in applied:
        print(f"[OK] {a}")
    for s in skipped:
        print(f"[SKIP] {s}")
    print(f"[OK] {len(applied)} status aplicados em {XLSX}")

    subprocess.run([sys.executable, os.path.join(BASE, "xlsx_to_json.py")], check=True)


if __name__ == "__main__":
    main()
