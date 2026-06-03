import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Backlog"

# ── cores ──
COR_HEADER   = "1F4E79"  # azul escuro
COR_EPIC     = "D6E4F7"
COR_FEATURE  = "E2EFDA"
COR_PBI      = "FFF2CC"
COR_TASK     = "FFFFFF"
COR_BUG      = "FCE4D6"

TIPO_COR = {
    "Epic":                 COR_EPIC,
    "Feature":              COR_FEATURE,
    "Product Backlog Item": COR_PBI,
    "Task":                 COR_TASK,
    "Bug":                  COR_BUG,
}

headers = [
    "csv_id", "tipo", "titulo", "descricao",
    "projeto_ado", "csv_id_pai", "prioridade",
    "sprint", "estimativa_horas", "tags", "ado_id"
]

col_widths = [8, 22, 45, 50, 15, 12, 12, 12, 18, 25, 10]

rows = [
    (1,  "Epic",                 "Dashboard Analítico Portal Único MS",                                            "", "SETDIG", "",  2, "",         "",  "portal-unico matomo", ""),
    (2,  "Feature",              "Painel de Acessos (Matomo)",                                                     "", "SETDIG", 1,   2, "Sprint 1", "",  "", ""),
    (3,  "Feature",              "Painel MS Digital (GA4)",                                                        "", "SETDIG", 1,   2, "Sprint 1", "",  "", ""),
    (4,  "Feature",              "Ranking Cartas de Serviço",                                                      "", "SETDIG", 1,   2, "Sprint 2", "",  "", ""),
    (5,  "Product Backlog Item", "Visualizar top páginas mais acessadas por período",                              "", "SETDIG", 2,   2, "Sprint 1", 8,   "", ""),
    (6,  "Product Backlog Item", "Exibir palavras-chave mais buscadas no portal",                                  "", "SETDIG", 2,   2, "Sprint 1", 6,   "", ""),
    (7,  "Product Backlog Item", "Mostrar jornada de navegação (transições) da carta de serviço principal",       "", "SETDIG", 2,   3, "Sprint 2", 13,  "", ""),
    (8,  "Product Backlog Item", "Integrar dados GA4 para app MS Digital",                                        "", "SETDIG", 3,   2, "Sprint 1", 16,  "", ""),
    (9,  "Product Backlog Item", "Filtrar e rankear Cartas de Serviço via regex nas URLs",                        "", "SETDIG", 4,   2, "Sprint 2", 8,   "", ""),
    (10, "Task",                 "Criar matomo_client.py com autenticação e builder de URL",                      "", "SETDIG", 5,   "", "Sprint 1", 4,   "", ""),
    (11, "Task",                 "Implementar Actions.getPageUrls no data_processor.py",                          "", "SETDIG", 5,   "", "Sprint 1", 3,   "", ""),
    (12, "Task",                 "Renderizar gráfico de barras no app.py (top 10 páginas)",                       "", "SETDIG", 5,   "", "Sprint 1", 2,   "", ""),
    (13, "Task",                 "Implementar Actions.getSiteSearchKeywords e exibir tabela",                     "", "SETDIG", 6,   "", "Sprint 1", 3,   "", ""),
    (14, "Task",                 "Implementar Transitions.getTransitionsForPageUrl",                              "", "SETDIG", 7,   "", "Sprint 2", 5,   "", ""),
    (15, "Task",                 "Exibir diagrama de fluxo (entrada/saída) com Plotly",                           "", "SETDIG", 7,   "", "Sprint 2", 4,   "", ""),
    (16, "Task",                 "Adicionar @st.cache_data para evitar timeout (period=year)",                    "", "SETDIG", 7,   "", "Sprint 2", 2,   "", ""),
    (17, "Task",                 "Conectar GA4 Data API no ga4_client.py (activeUsers + screenPageViews)",        "", "SETDIG", 8,   "", "Sprint 1", 6,   "", ""),
    (18, "Task",                 "Criar painel MS Digital no app.py com filtro de período",                       "", "SETDIG", 8,   "", "Sprint 1", 4,   "", ""),
    (19, "Task",                 "Definir regex pattern para Cartas de Serviço em data_processor.py",             "", "SETDIG", 9,   "", "Sprint 2", 3,   "", ""),
    (20, "Task",                 "Exibir ranking de cartas com acessos e comparativo mensal",                     "", "SETDIG", 9,   "", "Sprint 2", 3,   "", ""),
]

thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# cabeçalho
for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", start_color=COR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 22

# dados
for row_idx, row_data in enumerate(rows, start=2):
    tipo = row_data[1]
    cor = TIPO_COR.get(tipo, "FFFFFF")
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color=cor)
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.border = border
    ws.row_dimensions[row_idx].height = 18

# legenda — sheet separada
wl = wb.create_sheet("Legenda")
legend = [
    ("Tipo",                 "Cor",     "Descrição"),
    ("Epic",                 "#D6E4F7", "Agrupamento de alto nível"),
    ("Feature",              "#E2EFDA", "Entrega intermediária dentro do Epic"),
    ("Product Backlog Item", "#FFF2CC", "História de usuário implementável"),
    ("Task",                 "#FFFFFF", "Trabalho técnico atômico"),
    ("Bug",                  "#FCE4D6", "Correção de defeito"),
    ("", "", ""),
    ("Prioridade", "Valor", "Significado"),
    ("",           "1",     "Crítica"),
    ("",           "2",     "Alta"),
    ("",           "3",     "Média"),
    ("",           "4",     "Baixa"),
    ("", "", ""),
    ("ado_id", "", "Preencher após criação no Azure DevOps"),
    ("csv_id_pai", "", "Referência ao csv_id do item pai (neste arquivo)"),
]
for r, (a, b, c) in enumerate(legend, start=1):
    wl.cell(r, 1, a).font = Font(name="Arial", bold=(r in (1, 8)), size=10)
    wl.cell(r, 2, b).font = Font(name="Arial", size=10)
    wl.cell(r, 3, c).font = Font(name="Arial", size=10)
wl.column_dimensions["A"].width = 25
wl.column_dimensions["B"].width = 12
wl.column_dimensions["C"].width = 45

output = r"C:\Users\Fabio\Documents\SETDIG\2026\tfs\backlog.xlsx"
wb.save(output)
print(f"Salvo: {output}")
