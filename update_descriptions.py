"""Atualiza descrições no backlog.xlsx e regenera o HTML."""
import openpyxl, os, subprocess

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(BASE, "backlog.xlsx")

DESCRICOES = {
    1:  "Levantamento analítico do Portal Único de Serviços do MS, cruzando dados do Matomo (acessos), Google Analytics (app MS Digital) e banco de dados de cartas de serviço. Objetivo: subsidiar decisões da gestão sobre uso real dos serviços pelo cidadão.",
    2:  "Painel interativo de análise de acessos ao Portal Único via API do Matomo (idSite=298, webanalytics.ms.gov.br). Exibe páginas mais acessadas, buscas do cidadão e filtros por período (dia/semana/mês/ano).",
    3:  "Painel de análise do app MS Digital usando Google Analytics GA4. Exibe uso por categoria de serviço, comparando serviços nativos (screenPageViews) e redirecionados (eventCount/cliques). Pessoas = activeUsers; não somam entre categorias por sobreposição de usuários.",
    4:  "Identificação e ranqueamento automático das Cartas de Serviço mais acessadas no portal, usando regex para isolar slugs de serviço nas URLs retornadas pela API Matomo — excluindo diretórios raiz como /index e /busca.",
    5:  "Como gestor, quero visualizar as páginas mais acessadas do Portal Único em qualquer período (dia/semana/mês/ano), para identificar quais serviços têm maior demanda dos cidadãos e onde concentrar esforços de melhoria.",
    6:  "Como gestor, quero ver as palavras-chave mais buscadas na barra de pesquisa do portal, para identificar gargalos de navegação — situações em que o cidadão não encontra o serviço no menu e recorre à busca como alternativa.",
    7:  "Como gestor, quero ver de onde os usuários vieram antes de acessar o serviço mais popular e para onde foram depois (ou se abandonaram), para entender a jornada completa do cidadão no portal e identificar pontos de atrito.",
    8:  "Como gestor, quero visualizar os dados de uso do app MS Digital (pessoas únicas e acessos por categoria e serviço), para comparar o canal digital nativo com o portal web e identificar diferenças de adoção por tipo de serviço.",
    9:  "Como analista, quero que o sistema identifique automaticamente as Cartas de Serviço nas URLs do Matomo usando regex, gerando um ranking focado nos serviços reais oferecidos ao cidadão — separado das páginas institucionais e de navegação.",
    10: "Criar módulo api/matomo_client.py com: autenticação por token (leitura de .env/st.secrets), builder de URL base para requisições à API do Matomo (webanalytics.ms.gov.br, idSite=298), e função genérica get(method, params) que retorna JSON tratado.",
    11: "Implementar função em utils/data_processor.py que chama Actions.getPageUrls com parâmetro de período e retorna DataFrame pandas com colunas: label (URL), nb_visits, nb_uniq_visitors. Incluir filtro para remover URLs raiz (/index, /busca, /home).",
    12: "No app.py, adicionar seção 'Top Páginas' com st.selectbox para seleção de período e gráfico de barras horizontais Plotly mostrando as 10 URLs mais acessadas. Tooltip deve exibir nb_visits e nb_uniq_visitors.",
    13: "Implementar em data_processor.py chamada ao método Actions.getSiteSearchKeywords. Exibir resultado no app.py como st.dataframe com colunas: palavra-chave, total de buscas, percentual do total. Ordenar por frequência decrescente.",
    14: "Implementar em matomo_client.py chamada ao método Transitions.getTransitionsForPageUrl para a URL da carta de serviço mais acessada. Retornar estrutura com previousPages (de onde vieram), followingPages (para onde foram) e bounceRate.",
    15: "No app.py, renderizar Sankey diagram (Plotly go.Sankey) mostrando páginas de origem → carta de serviço → páginas de destino, usando os dados de transições. Limitar a top 5 origens e top 5 destinos para legibilidade.",
    16: "Decorar todas as funções que chamam a API Matomo com @st.cache_data(ttl=3600) no app.py. Testar especificamente com period=year para garantir que o tempo de carregamento não excede o timeout do servidor Matomo.",
    17: "Criar api/ga4_client.py com autenticação via service account (google-auth). Implementar função get_report(dimensions, metrics, start_date, end_date) usando GA4 Data API. Configurar property_id e credenciais no config.py / st.secrets.",
    18: "Adicionar aba ou seção 'MS Digital' no app.py com st.selectbox de período. Exibir tabela de categorias com colunas: categoria, pessoas (activeUsers), acessos (screenPageViews). Incluir botão de detalhe que abre modal com breakdown por serviço.",
    19: "Implementar função is_carta_servico(url) em utils/data_processor.py usando regex. Critérios positivos: slug com hífen, padrão de ação ('requerer', 'solicitar', 'agendar', 'pagar'). Critérios negativos: /index, /busca, /home, /imprimir isolado.",
    20: "No app.py, seção 'Cartas de Serviço': tabela com top 20 cartas filtradas mostrando URL, acessos do período e delta percentual vs. mês anterior. Destacar visualmente cartas com crescimento acima de 20% (badge ou cor de fundo).",
}

wb = openpyxl.load_workbook(XLSX)
ws = wb["Backlog"]
headers = [c.value for c in ws[1]]
desc_col = headers.index("descricao") + 1
id_col   = headers.index("csv_id") + 1

updated = 0
for row in ws.iter_rows(min_row=2):
    csv_id = row[id_col - 1].value
    if csv_id and int(csv_id) in DESCRICOES:
        row[desc_col - 1].value = DESCRICOES[int(csv_id)]
        updated += 1

import tempfile, shutil
tmp = XLSX + ".tmp"
wb.save(tmp)
try:
    shutil.move(tmp, XLSX)
    print(f"[OK] {updated} descrições atualizadas em backlog.xlsx")
except PermissionError:
    alt = os.path.join(BASE, "backlog_updated.xlsx")
    shutil.move(tmp, alt)
    print(f"[AVISO] backlog.xlsx está aberto no Excel. Salvo em: {alt}")
    print("        Feche o Excel, delete backlog.xlsx e renomeie backlog_updated.xlsx")
    import sys; sys.exit(0)

subprocess.run(["python", os.path.join(BASE, "xlsx_to_json.py")], check=True)
