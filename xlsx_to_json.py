"""
Converte backlog.xlsx → backlog.json + backlog_viewer.html (JSON inline).
Uso: python xlsx_to_json.py
"""
import json
import os
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(BASE, "backlog.xlsx")
JSON_OUT = os.path.join(BASE, "backlog.json")
HTML_OUT = os.path.join(BASE, "backlog_viewer.html")

STATUSES = ("A fazer", "Em progresso", "Finalizado")


def xlsx_to_items(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Backlog"]
    headers = [cell.value for cell in ws[1]]
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        item = dict(zip(headers, row))
        # normalizar
        item["csv_id"] = int(item["csv_id"]) if item["csv_id"] else None
        item["csv_id_pai"] = int(item["csv_id_pai"]) if item["csv_id_pai"] else None
        item["prioridade"] = int(item["prioridade"]) if item["prioridade"] else None
        item["estimativa_horas"] = int(item["estimativa_horas"]) if item["estimativa_horas"] else None
        item["ado_id"] = str(item["ado_id"]) if item["ado_id"] else ""
        item["descricao"] = str(item["descricao"]) if item["descricao"] else ""
        item["tags"] = str(item["tags"]) if item["tags"] else ""
        item["sprint"] = str(item["sprint"]) if item["sprint"] else ""
        st = str(item.get("status") or "").strip()
        item["status"] = st if st in STATUSES else "A fazer"
        items.append(item)
    return items


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Backlog — SETDIG</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
:root {
  --bg:        #F8F9FB;
  --surface:   #FFFFFF;
  --border:    #E5E8ED;
  --border-2:  #D0D5DE;
  --text:      #1A1D23;
  --text-sub:  #5A6070;
  --text-muted:#9299A8;
  --selected:  #EEF3FD;
  --sel-bar:   #0078D4;
  --hover:     #F3F5F9;

  /* TFS / Azure DevOps official colors */
  --c-epic:    #FF7B00;
  --c-feat:    #773B93;
  --c-pbi:     #009CCC;
  --c-task:    #F2CB1D;
  --c-bug:     #CC293D;

  --c-epic-bg: #FFF3E0;
  --c-feat-bg: #EDE7F6;
  --c-pbi-bg:  #E1F5FE;
  --c-task-bg: #FFF8E1;
  --c-bug-bg:  #FDECEA;

  /* Status */
  --s-todo:    #9299A8;
  --s-prog:    #0078D4;
  --s-done:    #107C10;

  --indent: 20px;
  --row-h: 32px;
}

*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

body {
  font-family: 'Plus Jakarta Sans', 'Segoe UI', system-ui, sans-serif;
  background: var(--bg);
  color: var(--text);
  height: 100vh;
  display: flex;
  flex-direction: column;
  overflow: hidden;
  font-size: 13px;
}

/* ── HEADER ── */
header {
  display: flex;
  align-items: center;
  gap: 0;
  height: 48px;
  background: #fff;
  border-bottom: 1px solid var(--border);
  flex-shrink: 0;
  padding: 0 20px;
}

.hdr-brand {
  display: flex;
  align-items: center;
  gap: 10px;
  padding-right: 20px;
  border-right: 1px solid var(--border);
  margin-right: 16px;
}

.hdr-brand-icon {
  width: 28px; height: 28px;
  background: #0078D4;
  border-radius: 6px;
  display: flex; align-items: center; justify-content: center;
}

.hdr-brand-icon svg { width: 16px; height: 16px; }

.hdr-org {
  font-size: 13px;
  font-weight: 600;
  color: var(--text);
  line-height: 1;
}

.hdr-team {
  font-size: 11px;
  color: var(--text-muted);
  margin-top: 2px;
}

.hdr-title {
  font-size: 13px;
  font-weight: 500;
  color: var(--text-sub);
}

.hdr-right {
  margin-left: auto;
  display: flex;
  align-items: center;
  gap: 8px;
  font-size: 11px;
  color: var(--text-muted);
  font-family: 'JetBrains Mono', monospace;
}

.hdr-pill {
  background: var(--bg);
  border: 1px solid var(--border);
  border-radius: 10px;
  padding: 2px 8px;
  font-size: 10px;
}

/* ── LAYOUT ── */
.layout { display: flex; flex: 1; overflow: hidden; }

/* ── TREE ── */
.tree-panel {
  width: 50%;
  min-width: 320px;
  overflow-y: auto;
  background: #fff;
  border-right: 1px solid var(--border);
  padding: 0 0 32px;
  scrollbar-width: thin;
  scrollbar-color: var(--border-2) transparent;
}
#tree { padding-top: 8px; }
.tree-panel::-webkit-scrollbar { width: 5px; }
.tree-panel::-webkit-scrollbar-thumb { background: var(--border-2); border-radius: 3px; }

.node { position: relative; }

.node-row {
  display: flex;
  align-items: center;
  height: var(--row-h);
  gap: 6px;
  cursor: pointer;
  border-left: 2px solid transparent;
  padding-right: 12px;
  transition: background 100ms, border-color 100ms;
  position: relative;
}
.node-row:hover { background: var(--hover); }
.node-row.selected {
  background: var(--selected);
  border-left-color: var(--sel-bar);
}

/* ── BOWTIE ICON ── */
.wi-icon {
  width: 16px; height: 16px;
  border-radius: 3px;
  display: flex; align-items: center; justify-content: center;
  flex-shrink: 0;
}
.wi-icon svg { width: 10px; height: 10px; }

/* toggle chevron */
.toggle {
  width: 16px; height: 16px;
  display: flex; align-items: center; justify-content: center;
  flex-shrink: 0;
  color: var(--text-muted);
  font-size: 8px;
  transition: transform 140ms ease;
  border-radius: 3px;
}
.toggle:hover { background: var(--border); }
.toggle.open { transform: rotate(90deg); }
.toggle.leaf { opacity: 0; pointer-events: none; }

.title-text {
  flex: 1;
  overflow: hidden;
  text-overflow: ellipsis;
  white-space: nowrap;
  font-size: 13px;
  color: var(--text);
  font-weight: 400;
}
.node-row.selected .title-text { font-weight: 500; color: #0053a0; }

/* guide line */
.children-wrap { position: relative; }
.children-wrap::before {
  content: '';
  position: absolute;
  top: 0; bottom: 8px;
  width: 1px;
  background: var(--border);
  left: var(--guide-x);
  pointer-events: none;
}

.children { display: none; }
.children.open { display: block; }

/* ── DETAIL PANEL ── */
.detail-panel {
  flex: 1;
  background: var(--bg);
  overflow-y: auto;
  padding: 0;
  scrollbar-width: thin;
  scrollbar-color: var(--border-2) transparent;
}
.detail-panel::-webkit-scrollbar { width: 5px; }
.detail-panel::-webkit-scrollbar-thumb { background: var(--border-2); border-radius: 3px; }

.empty-state {
  display: flex; flex-direction: column;
  align-items: center; justify-content: center;
  height: 100%; gap: 10px; color: var(--text-muted);
}
.empty-icon { font-size: 28px; opacity: .25; }
.empty-msg { font-size: 12px; }

/* detail header bar */
.det-header {
  display: flex;
  align-items: center;
  gap: 10px;
  padding: 16px 28px;
  background: #fff;
  border-bottom: 1px solid var(--border);
}

.det-wi-icon {
  width: 24px; height: 24px;
  border-radius: 4px;
  display: flex; align-items: center; justify-content: center;
  flex-shrink: 0;
}
.det-wi-icon svg { width: 14px; height: 14px; }

.det-type-label {
  font-size: 11px;
  font-weight: 600;
  letter-spacing: .04em;
  text-transform: uppercase;
}

.det-sep { width: 1px; height: 16px; background: var(--border); }

.det-id {
  font-family: 'JetBrains Mono', monospace;
  font-size: 11px;
  color: var(--text-muted);
}

/* detail body */
.det-body {
  padding: 24px 28px 40px;
}

.det-title {
  font-size: 20px;
  font-weight: 700;
  color: var(--text);
  line-height: 1.35;
  margin-bottom: 24px;
}

/* field */
.field { margin-bottom: 20px; }

.field-top {
  display: flex;
  align-items: center;
  justify-content: space-between;
  margin-bottom: 6px;
}

.field-label {
  font-size: 10px;
  font-weight: 700;
  letter-spacing: .08em;
  text-transform: uppercase;
  color: var(--text-muted);
}

.copy-btn {
  font-size: 11px;
  font-family: 'Plus Jakarta Sans', sans-serif;
  font-weight: 500;
  padding: 3px 10px;
  border: 1px solid var(--border-2);
  border-radius: 4px;
  background: #fff;
  color: var(--text-sub);
  cursor: pointer;
  transition: all 100ms;
}
.copy-btn:hover { background: #EEF3FD; border-color: #0078D4; color: #0078D4; }
.copy-btn.ok { background: #E6F4EA; border-color: #1E7E34; color: #1E7E34; }

.field-value {
  background: #fff;
  border: 1px solid var(--border);
  border-radius: 4px;
  padding: 10px 14px;
  font-size: 13px;
  line-height: 1.6;
  color: var(--text);
  white-space: pre-wrap;
  word-break: break-word;
  min-height: 38px;
}
.field-value.title-v { font-weight: 500; font-size: 14px; }
.field-value.empty-v { color: var(--text-muted); font-style: italic; font-size: 12px; }

/* ── STATUS ── */
.st-dot {
  width: 8px; height: 8px;
  border-radius: 50%;
  flex-shrink: 0;
  transition: background 150ms;
}
.node-prog {
  font-family: 'JetBrains Mono', monospace;
  font-size: 10px;
  color: var(--text-muted);
  flex-shrink: 0;
}
.node-row.st-done .title-text { opacity: .55; }

/* filter bar */
.filter-bar {
  display: flex;
  gap: 6px;
  padding: 10px 12px;
  position: sticky;
  top: 0;
  z-index: 3;
  background: #fff;
  border-bottom: 1px solid var(--border);
}
.chip {
  font-family: 'Plus Jakarta Sans', sans-serif;
  font-size: 11px;
  font-weight: 600;
  padding: 4px 12px;
  border-radius: 12px;
  border: 1px solid var(--border-2);
  background: #fff;
  color: var(--text-sub);
  cursor: pointer;
  display: inline-flex;
  align-items: center;
  gap: 6px;
  transition: all 120ms;
}
.chip .dot { width: 7px; height: 7px; border-radius: 50%; }
.chip:hover { border-color: #0078D4; color: #0078D4; }
.chip.active { background: var(--text); border-color: var(--text); color: #fff; }

/* header stats + progress */
.hdr-stat {
  display: inline-flex;
  align-items: center;
  gap: 5px;
  font-size: 10px;
}
.hdr-stat .dot { width: 7px; height: 7px; border-radius: 50%; }
.hdr-progress {
  width: 110px; height: 5px;
  background: var(--border);
  border-radius: 3px;
  overflow: hidden;
}
.hdr-progress-fill {
  height: 100%; width: 0%;
  background: var(--s-done);
  border-radius: 3px;
  transition: width 250ms ease;
}
.hdr-pct {
  font-size: 10px;
  font-weight: 600;
  color: var(--text-sub);
  min-width: 30px;
}
.export-btn {
  font-family: 'Plus Jakarta Sans', sans-serif;
  font-size: 11px;
  font-weight: 600;
  padding: 4px 12px;
  border: 1px solid var(--border-2);
  border-radius: 4px;
  background: #fff;
  color: var(--text-sub);
  cursor: pointer;
  transition: all 100ms;
}
.export-btn:hover { background: #EEF3FD; border-color: #0078D4; color: #0078D4; }

/* segmented status control (detail) */
.status-seg {
  display: inline-flex;
  margin-left: auto;
  border: 1px solid var(--border-2);
  border-radius: 6px;
  overflow: hidden;
  flex-shrink: 0;
}
.status-seg button {
  font-family: 'Plus Jakarta Sans', sans-serif;
  font-size: 11px;
  font-weight: 600;
  padding: 5px 12px;
  background: #fff;
  color: var(--text-sub);
  border: none;
  border-right: 1px solid var(--border);
  cursor: pointer;
  display: inline-flex;
  align-items: center;
  gap: 6px;
  transition: all 120ms;
}
.status-seg button:last-child { border-right: none; }
.status-seg button .dot { width: 7px; height: 7px; border-radius: 50%; }
.status-seg button:hover { background: var(--hover); }
.status-seg button.active { color: #fff; }
.status-seg button.active.todo { background: var(--s-todo); }
.status-seg button.active.prog { background: var(--s-prog); }
.status-seg button.active.done { background: var(--s-done); }
.status-seg button.active .dot { background: #fff !important; }
</style>
</head>
<body>

<header>
  <div class="hdr-brand">
    <div class="hdr-brand-icon">
      <svg viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg">
        <path d="M1 1h6v6H1zM9 1h6v6H9zM1 9h6v6H1zM9 9h6v6H9z" fill="white" opacity=".9"/>
      </svg>
    </div>
    <div>
      <div class="hdr-org">SETDIG</div>
      <div class="hdr-team">SGD — EDS Team</div>
    </div>
  </div>
  <span class="hdr-title">Backlog · Dashboard Analítico Portal Único MS</span>
  <div class="hdr-right">
    <span class="hdr-stat"><span class="dot" style="background:var(--s-todo)"></span><span id="cnt-todo">0</span></span>
    <span class="hdr-stat"><span class="dot" style="background:var(--s-prog)"></span><span id="cnt-prog">0</span></span>
    <span class="hdr-stat"><span class="dot" style="background:var(--s-done)"></span><span id="cnt-done">0</span></span>
    <div class="hdr-progress"><div class="hdr-progress-fill" id="hdr-fill"></div></div>
    <span class="hdr-pct" id="hdr-pct">0%</span>
    <button class="export-btn" id="export-btn" title="Baixa status_updates.json — aplique no xlsx com: python apply_status.py">Exportar status</button>
    <span class="hdr-pill" id="hdr-count">— items</span>
  </div>
</header>

<div class="layout">
  <div class="tree-panel">
    <div class="filter-bar" id="filterbar"></div>
    <div id="tree"></div>
  </div>
  <div class="detail-panel" id="detail">
    <div class="empty-state">
      <div class="empty-icon">⊞</div>
      <div class="empty-msg">Selecione um work item</div>
    </div>
  </div>
</div>

<script>
const DATA = __JSON_DATA__;

// ── Bowtie SVG icons (ADO style) ──────────────────────────────────────────
const ICONS = {
  "Epic": {
    bg: "#FF7B00", fg: "#FFF3E0",
    svg: `<svg viewBox="0 0 10 10" fill="white" xmlns="http://www.w3.org/2000/svg">
      <path d="M1 3.2L2.5 6 5 2.2 7.5 6 9 3.2V8H1z"/>
      <circle cx="1" cy="2.6" r="0.6"/>
      <circle cx="5" cy="1.6" r="0.6"/>
      <circle cx="9" cy="2.6" r="0.6"/>
    </svg>`
  },
  "Feature": {
    bg: "#773B93", fg: "#EDE7F6",
    svg: `<svg viewBox="0 0 10 10" fill="white" xmlns="http://www.w3.org/2000/svg">
      <path d="M3 1.2h4v3a2 2 0 0 1-4 0V1.2z"/>
      <rect x="4.3" y="5.6" width="1.4" height="1.6"/>
      <rect x="2.8" y="7.6" width="4.4" height="1.2" rx="0.2"/>
      <path d="M2 2.4H1v1.4a1.2 1.2 0 0 0 1.2 1.2H3" fill="none" stroke="white" stroke-width="0.7"/>
      <path d="M8 2.4h1v1.4a1.2 1.2 0 0 1-1.2 1.2H7" fill="none" stroke="white" stroke-width="0.7"/>
    </svg>`
  },
  "Product Backlog Item": {
    bg: "#009CCC", fg: "#E1F5FE",
    svg: `<svg viewBox="0 0 10 10" fill="none" stroke="white" stroke-width="0.9" xmlns="http://www.w3.org/2000/svg">
      <rect x="1.2" y="1.5" width="7.6" height="7" rx="0.6" fill="white" fill-opacity="0.15"/>
      <line x1="2.4" y1="3.4" x2="7.6" y2="3.4"/>
      <line x1="2.4" y1="5"   x2="7.6" y2="5"/>
      <line x1="2.4" y1="6.6" x2="5.6" y2="6.6"/>
    </svg>`
  },
  "Task": {
    bg: "#F2CB1D", fg: "#FFF8E1",
    svg: `<svg viewBox="0 0 10 10" fill="none" stroke="white" stroke-width="1.4" stroke-linecap="round" stroke-linejoin="round" xmlns="http://www.w3.org/2000/svg">
      <rect x="1" y="1" width="8" height="8" rx="1" fill="white" fill-opacity="0.15" stroke="white" stroke-width="0.9"/>
      <polyline points="3,5 4.3,6.5 7,3.5"/>
    </svg>`
  },
  "Bug": {
    bg: "#CC293D", fg: "#FDECEA",
    svg: `<svg viewBox="0 0 10 10" fill="white" xmlns="http://www.w3.org/2000/svg">
      <ellipse cx="5" cy="5.5" rx="2.5" ry="3"/>
      <path d="M3.5 2.5 Q5 1 6.5 2.5" stroke="white" stroke-width="1" fill="none"/>
      <line x1="1" y1="4" x2="3" y2="4.5" stroke="white" stroke-width="1"/>
      <line x1="9" y1="4" x2="7" y2="4.5" stroke="white" stroke-width="1"/>
      <line x1="1" y1="6.5" x2="2.5" y2="6.5" stroke="white" stroke-width="1"/>
      <line x1="9" y1="6.5" x2="7.5" y2="6.5" stroke="white" stroke-width="1"/>
    </svg>`
  },
};
const LABEL = {
  "Epic": "Epic", "Feature": "Feature",
  "Product Backlog Item": "Product Backlog Item",
  "Task": "Task", "Bug": "Bug",
};
const INDENT_PX = 20;

const map = {};
DATA.items.forEach(it => { map[it.csv_id] = { ...it, children: [], parent: null }; });
const roots = [];
DATA.items.forEach(it => {
  const n = map[it.csv_id];
  if (it.csv_id_pai && map[it.csv_id_pai]) {
    map[it.csv_id_pai].children.push(n);
    n.parent = map[it.csv_id_pai];
  } else roots.push(n);
});

document.getElementById("hdr-count").textContent =
  `${DATA.items.length} items`;

// ── Status (localStorage overrides) ───────────────────────────────────────
const STATUSES = ["A fazer", "Em progresso", "Finalizado"];
const ST_CLS = { "A fazer": "todo", "Em progresso": "prog", "Finalizado": "done" };
const ST_VAR = { "A fazer": "var(--s-todo)", "Em progresso": "var(--s-prog)", "Finalizado": "var(--s-done)" };
const LS_KEY = "tfs-backlog-status";

let overrides = {};
try { overrides = JSON.parse(localStorage.getItem(LS_KEY) || "{}"); } catch (e) { overrides = {}; }
// poda overrides redundantes (já sincronizados no xlsx) ou inválidos
Object.keys(overrides).forEach(id => {
  if (!map[id] || !STATUSES.includes(overrides[id]) ||
      overrides[id] === (map[id].status || "A fazer")) delete overrides[id];
});
localStorage.setItem(LS_KEY, JSON.stringify(overrides));

function effStatus(n) { return overrides[n.csv_id] || n.status || "A fazer"; }

function descendants(n) {
  let out = [];
  n.children.forEach(c => { out.push(c); out = out.concat(descendants(c)); });
  return out;
}

function refreshNode(n) {
  const st = effStatus(n);
  if (n.dotEl) n.dotEl.style.background = ST_VAR[st];
  if (n.rowEl) n.rowEl.classList.toggle("st-done", st === "Finalizado");
  if (n.progEl) {
    const ds = descendants(n);
    const done = ds.filter(d => effStatus(d) === "Finalizado").length;
    n.progEl.textContent = `${done}/${ds.length}`;
  }
}

function updateHeader() {
  const c = { "A fazer": 0, "Em progresso": 0, "Finalizado": 0 };
  DATA.items.forEach(it => c[effStatus(map[it.csv_id])]++);
  document.getElementById("cnt-todo").textContent = c["A fazer"];
  document.getElementById("cnt-prog").textContent = c["Em progresso"];
  document.getElementById("cnt-done").textContent = c["Finalizado"];
  const total = DATA.items.length;
  const pct = total ? Math.round(100 * c["Finalizado"] / total) : 0;
  document.getElementById("hdr-fill").style.width = pct + "%";
  document.getElementById("hdr-pct").textContent = pct + "%";
}

let currentSeg = null;
function syncSeg(n) {
  if (!currentSeg || currentSeg.node !== n) return;
  const st = effStatus(n);
  currentSeg.seg.querySelectorAll("button").forEach((b, i) => {
    b.classList.toggle("active", STATUSES[i] === st);
  });
}

function setStatus(n, st) {
  const base = n.status || "A fazer";
  if (st === base) delete overrides[n.csv_id];
  else overrides[n.csv_id] = st;
  localStorage.setItem(LS_KEY, JSON.stringify(overrides));
  let p = n;
  while (p) { refreshNode(p); p = p.parent; }
  updateHeader();
  applyFilter();
  syncSeg(n);
}

// ── Filtro por status ─────────────────────────────────────────────────────
let filter = "Todos";

function nodeMatches(n) {
  if (filter === "Todos") return true;
  if (effStatus(n) === filter) return true;
  return n.children.some(nodeMatches);
}

function applyFilter() {
  Object.values(map).forEach(n => {
    const show = nodeMatches(n);
    if (n.wrapEl) n.wrapEl.style.display = show ? "" : "none";
    if (filter !== "Todos" && show && n.children.length && n.children.some(nodeMatches)) {
      n.childrenEl.classList.add("open");
      n.toggleEl.classList.add("open");
    }
  });
}

function wiIcon(tipo, size = 16) {
  const ic = ICONS[tipo] || ICONS["Task"];
  const el = document.createElement("div");
  el.className = "wi-icon";
  el.style.cssText = `width:${size}px;height:${size}px;background:${ic.bg};border-radius:${size<=16?'3':'4'}px;display:flex;align-items:center;justify-content:center;flex-shrink:0`;
  el.innerHTML = ic.svg;
  return el;
}

function renderNode(node, depth) {
  const wrap = document.createElement("div");
  wrap.className = "node";

  const row = document.createElement("div");
  row.className = "node-row";
  row.style.paddingLeft = (8 + depth * INDENT_PX) + "px";
  row.dataset.id = node.csv_id;

  const toggle = document.createElement("span");
  toggle.className = "toggle" + (node.children.length ? "" : " leaf");
  toggle.innerHTML = "▶";

  const icon = wiIcon(node.tipo, 16);

  const title = document.createElement("span");
  title.className = "title-text";
  title.textContent = node.titulo;

  row.appendChild(toggle);
  row.appendChild(icon);
  row.appendChild(title);

  if (node.children.length) {
    const prog = document.createElement("span");
    prog.className = "node-prog";
    node.progEl = prog;
    row.appendChild(prog);
  }

  const dot = document.createElement("span");
  dot.className = "st-dot";
  row.appendChild(dot);

  node.rowEl = row;
  node.dotEl = dot;
  node.wrapEl = wrap;
  node.toggleEl = toggle;

  // children
  const childWrap = document.createElement("div");
  childWrap.className = "children-wrap";
  childWrap.style.setProperty("--guide-x", (16 + depth * INDENT_PX) + "px");

  const childrenDiv = document.createElement("div");
  childrenDiv.className = "children";
  node.children.forEach(c => childrenDiv.appendChild(renderNode(c, depth + 1)));
  childWrap.appendChild(childrenDiv);
  node.childrenEl = childrenDiv;

  row.addEventListener("click", e => {
    e.stopPropagation();
    if (node.children.length) {
      const open = childrenDiv.classList.toggle("open");
      toggle.classList.toggle("open", open);
    }
    document.querySelectorAll(".node-row.selected").forEach(r => r.classList.remove("selected"));
    row.classList.add("selected");
    renderDetail(node);
  });

  wrap.appendChild(row);
  wrap.appendChild(childWrap);
  return wrap;
}

function copy(text, btn) {
  (navigator.clipboard?.writeText(text) || Promise.reject())
    .then(() => flash(btn))
    .catch(() => {
      const ta = Object.assign(document.createElement("textarea"), { value: text });
      document.body.appendChild(ta); ta.select();
      document.execCommand("copy"); document.body.removeChild(ta);
      flash(btn);
    });
}
function flash(btn) {
  const t = btn.textContent;
  btn.textContent = "✓ Copiado"; btn.classList.add("ok");
  setTimeout(() => { btn.textContent = t; btn.classList.remove("ok"); }, 1400);
}

function field(label, value, cls = "") {
  const w = document.createElement("div"); w.className = "field";
  const top = document.createElement("div"); top.className = "field-top";
  const lbl = document.createElement("span"); lbl.className = "field-label"; lbl.textContent = label;
  const btn = document.createElement("button"); btn.className = "copy-btn"; btn.textContent = "Copiar";
  btn.addEventListener("click", () => copy(value || "", btn));
  top.appendChild(lbl); top.appendChild(btn);
  const val = document.createElement("div");
  val.className = "field-value" + (cls ? " "+cls : "") + (value ? "" : " empty-v");
  val.textContent = value || "(sem preenchimento)";
  w.appendChild(top); w.appendChild(val);
  return w;
}

function renderDetail(node) {
  const panel = document.getElementById("detail");
  panel.innerHTML = "";

  const ic = ICONS[node.tipo] || ICONS["Task"];

  // header bar
  const hdr = document.createElement("div"); hdr.className = "det-header";
  const bigIcon = wiIcon(node.tipo, 24); bigIcon.className = "det-wi-icon";
  bigIcon.style.cssText = `width:24px;height:24px;background:${ic.bg};border-radius:4px;display:flex;align-items:center;justify-content:center;flex-shrink:0`;
  const typeLabel = document.createElement("span"); typeLabel.className = "det-type-label";
  typeLabel.style.color = ic.bg; typeLabel.textContent = LABEL[node.tipo] || node.tipo;
  const sep = document.createElement("div"); sep.className = "det-sep";
  const idEl = document.createElement("span"); idEl.className = "det-id";
  idEl.textContent = node.ado_id ? `#${node.ado_id}` : `CSV-${node.csv_id}`;
  hdr.appendChild(bigIcon); hdr.appendChild(typeLabel);
  hdr.appendChild(sep); hdr.appendChild(idEl);

  // status segmented control
  const seg = document.createElement("div"); seg.className = "status-seg";
  STATUSES.forEach(st => {
    const b = document.createElement("button");
    b.classList.add(ST_CLS[st]);
    const d = document.createElement("span"); d.className = "dot";
    d.style.background = ST_VAR[st];
    b.appendChild(d);
    b.appendChild(document.createTextNode(st));
    b.addEventListener("click", () => setStatus(node, st));
    seg.appendChild(b);
  });
  hdr.appendChild(seg);
  currentSeg = { node, seg };
  syncSeg(node);

  panel.appendChild(hdr);

  // body
  const body = document.createElement("div"); body.className = "det-body";
  const titleEl = document.createElement("div"); titleEl.className = "det-title";
  titleEl.textContent = node.titulo;
  body.appendChild(titleEl);
  body.appendChild(field("Título", node.titulo, "title-v"));
  body.appendChild(field("Descrição", node.descricao));
  panel.appendChild(body);
}

const treeEl = document.getElementById("tree");
roots.forEach(r => treeEl.appendChild(renderNode(r, 0)));

// filter chips
const fb = document.getElementById("filterbar");
["Todos", ...STATUSES].forEach(f => {
  const b = document.createElement("button");
  b.className = "chip" + (f === filter ? " active" : "");
  if (f !== "Todos") {
    const d = document.createElement("span"); d.className = "dot";
    d.style.background = ST_VAR[f];
    b.appendChild(d);
  }
  b.appendChild(document.createTextNode(f));
  b.addEventListener("click", () => {
    filter = f;
    fb.querySelectorAll(".chip").forEach(c => c.classList.remove("active"));
    b.classList.add("active");
    applyFilter();
  });
  fb.appendChild(b);
});

// export status overrides
document.getElementById("export-btn").addEventListener("click", () => {
  const blob = new Blob([JSON.stringify(overrides, null, 2)], { type: "application/json" });
  const a = document.createElement("a");
  a.href = URL.createObjectURL(blob);
  a.download = "status_updates.json";
  a.click();
  URL.revokeObjectURL(a.href);
});

Object.values(map).forEach(refreshNode);
updateHeader();

document.querySelectorAll(".node-row[data-id]").forEach(row => {
  const n = map[parseInt(row.dataset.id)];
  if (n?.tipo === "Epic" && n.children.length) row.click();
});
</script>
</body>
</html>
"""


def generate(xlsx_path=XLSX, json_out=JSON_OUT, html_out=HTML_OUT):
    items = xlsx_to_items(xlsx_path)
    data = {"items": items}

    with open(json_out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"[OK] {json_out} — {len(items)} itens")

    json_inline = json.dumps(data, ensure_ascii=False)
    html = HTML_TEMPLATE.replace("__JSON_DATA__", json_inline)
    with open(html_out, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[OK] {html_out}")


if __name__ == "__main__":
    generate()
